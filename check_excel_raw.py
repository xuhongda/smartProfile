import openpyxl

# 读取 Excel 文件内容
test_files = [
    'backend/tests/八17期末下.xlsx',
    'backend/tests/成绩单.xlsx'
]

for file_path in test_files:
    print(f'检查文件: {file_path}')
    try:
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            print(f'  工作表: {sheet_name}')
            ws = wb[sheet_name]
            # 遍历所有单元格
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and '物理' in cell:
                        print(f'    找到 "物理": {cell}')
    except Exception as e:
        print(f'  读取失败: {str(e)}')
    print('---')
